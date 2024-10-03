# Importa a função load_workbook da biblioteca openpyxl
# openpyxl é usada para ler e modificar arquivos Excel no formato .xlsx (visualizar em: https://openpyxl.readthedocs.io/en/stable/)
# load_workbook: Essa função carrega uma planilha Excel existente para que possa ser manipulada no Python
from openpyxl import load_workbook

# Carrega o arquivo Excel
# 'workbook': Representa todo o arquivo Excel, permitindo acessar e modificar as diferentes planilhas dentro dele.
# Carrega o arquivo Excel chamado "dados_ficticios_1000_linhas.xlsx" e armazena o conteúdo em um objeto chamado workbook.
workbook = load_workbook(filename="Arquivo01_teste.xlsx")

# Acessa a primeira planilha ativa do arquivo Excel carregado
# Se houver mais de uma planilha, sheet representa a primeira por padrão (a ativa).
sheet = workbook.active

# Ler dados de uma célula específica
# Aqui, acessamos o valor da célula "A1", que fica na coluna A, linha 1.
valor = sheet["A1"].value

# Imprime o valor da célula A1, mostrando o conteúdo da célula na tela.
print(f"Valor de A1: {valor}")

# Escrever um novo valor na célula B1
# Aqui, modificamos o conteúdo da célula "B1", atribuindo a string "Novo Valor" a ela.
sheet["B1"] = "Novo Valor"

# Salvar as alterações feitas na planilha
# Como alteramos o valor de uma célula, salvamos o arquivo Excel com um novo nome "dados_atualizados.xlsx"
# Isso garante que o arquivo original "dados_ficticios_1000_linhas.xlsx" não seja sobrescrito.
workbook.save("Arquivo01_teste_atualizado.xlsx")
